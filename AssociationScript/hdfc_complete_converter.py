#!/usr/bin/env python3
"""
Complete HDFC Bank Statement Converter
Extract both credits AND debits with flat number detection
Creates separate Excel sheets for Credits and Debits
Supports both PDF and Excel input formats
"""

import sys
import os
from pathlib import Path
import re
import pandas as pd
from datetime import datetime
import json

def select_input_source():
    """Let user choose between PDF or Excel input"""
    print("🔍 SELECT INPUT SOURCE")
    print("=" * 40)
    
    input_dir = Path("input")
    pdf_files = list(input_dir.glob("*.pdf"))
    excel_files = list(input_dir.glob("*.xls*"))
    
    print(f"� Found {len(pdf_files)} PDF files and {len(excel_files)} Excel files")
    print()
    
    if pdf_files:
        print("PDF Files:")
        for idx, pdf in enumerate(pdf_files, 1):
            print(f"  {idx}. {pdf.name}")
    
    if excel_files:
        print("\nExcel Files:")
        for idx, excel in enumerate(excel_files, 1):
            print(f"  {idx}. {excel.name}")
    
    print()
    
    if not pdf_files and not excel_files:
        print("❌ No PDF or Excel files found in /input/ directory")
        return None, None
    
    # Prompt user for choice
    while True:
        print("\nChoose input type:")
        print("  1. PDF")
        print("  2. Excel")
        choice = input("Enter your choice (1 or 2): ").strip()
        
        if choice == "1" and pdf_files:
            if len(pdf_files) == 1:
                return "pdf", pdf_files[0]
            else:
                print("\nSelect PDF file:")
                for idx, pdf in enumerate(pdf_files, 1):
                    print(f"  {idx}. {pdf.name}")
                file_choice = input("Enter file number: ").strip()
                try:
                    file_idx = int(file_choice) - 1
                    if 0 <= file_idx < len(pdf_files):
                        return "pdf", pdf_files[file_idx]
                except ValueError:
                    pass
                print("❌ Invalid choice")
        
        elif choice == "2" and excel_files:
            if len(excel_files) == 1:
                return "excel", excel_files[0]
            else:
                print("\nSelect Excel file:")
                for idx, excel in enumerate(excel_files, 1):
                    print(f"  {idx}. {excel.name}")
                file_choice = input("Enter file number: ").strip()
                try:
                    file_idx = int(file_choice) - 1
                    if 0 <= file_idx < len(excel_files):
                        return "excel", excel_files[file_idx]
                except ValueError:
                    pass
                print("❌ Invalid choice")
        else:
            print("❌ Invalid choice or no files of that type available")

def extract_from_pdf(pdf_file):
    """Extract text from HDFC bank statement PDF"""
    print(f"📄 Extracting text from PDF: {pdf_file.name}")
    print("-" * 40)
    
    try:
        import pdfplumber
        
        full_text = ""
        with pdfplumber.open(pdf_file) as pdf:
            print(f"📄 Extracting from {len(pdf.pages)} pages...")
            
            for page_num, page in enumerate(pdf.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    full_text += page_text + "\n"
                if page_num % 5 == 0:
                    print(f"   ✅ Processed {page_num} pages")
        
        print(f"✅ PDF extraction complete ({len(full_text)} characters)")
        return full_text
        
    except Exception as e:
        print(f"❌ Error extracting PDF: {e}")
        return None

def extract_from_excel(excel_file):
    """Extract transaction text from Excel file"""
    print(f"📊 Extracting data from Excel: {excel_file.name}")
    print("-" * 40)
    
    try:
        # Read Excel file without headers first to find where data starts
        df_raw = pd.read_excel(excel_file, header=None)
        
        print(f"📊 Found {len(df_raw)} rows and {len(df_raw.columns)} columns")
        
        # Find the header row (usually contains "Date", "Narration", etc.)
        header_row = None
        for idx, row in df_raw.iterrows():
            row_text = ' '.join([str(x).lower() for x in row.values if pd.notna(x)])
            if 'date' in row_text and 'narration' in row_text:
                header_row = idx
                print(f"✅ Found header at row {idx}")
                break
        
        if header_row is None:
            print("❌ Could not find transaction headers in Excel file")
            return None, None
        
        # Re-read with proper header row
        df = pd.read_excel(excel_file, header=header_row)
        
        print(f"📊 Columns detected: {list(df.columns)}")
        
        # Filter out separator rows (those starting with ***)
        df = df[~df.iloc[:, 0].astype(str).str.startswith('*', na=False)]
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        print(f"📊 Total transaction rows after cleanup: {len(df)}")
        
        return df, df
        
    except Exception as e:
        print(f"❌ Error extracting Excel: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def extract_flat_number(description):
    """Extract flat number from transaction description"""
    # Enhanced regex patterns for flat numbers
    patterns = [
        # Direct patterns: A001, B123, C456
        r'\b([ABC])(\d{3})\b',
        # Patterns with spaces: A 001, B 123
        r'\b([ABC])\s*(\d{3})\b',
        # Embedded patterns: A001ABC, XYZB123DEF
        r'([ABC])(\d{3})(?=[A-Z]|$|\s)',
        # Patterns after keywords
        r'(?:FLAT|UNIT|APT|APARTMENT)\s*[:\-]?\s*([ABC])(\d{3})',
        # Patterns before keywords  
        r'([ABC])(\d{3})\s*(?:FLAT|UNIT|APT|APARTMENT)',
    ]
    
    description_upper = description.upper()
    
    for pattern in patterns:
        matches = re.findall(pattern, description_upper)
        for match in matches:
            if len(match) == 2:  # (letter, number)
                letter, number = match
                num_int = int(number)
                
                # Validate ranges
                if ((letter == 'A' and 1 <= num_int <= 320) or
                    (letter == 'B' and 1 <= num_int <= 312) or
                    (letter == 'C' and 1 <= num_int <= 318)):
                    return f"{letter}{number}"
    
    return None

def parse_transactions(text, input_type="pdf", df=None):
    """Parse all transactions from extracted text with balance-based credit/debit detection"""
    print("\n💰 STEP 2: Parsing Transactions")
    print("=" * 40)
    
    transactions = []
    previous_balance = None
    
    # If Excel input, parse from dataframe directly
    if input_type == "excel" and df is not None:
        return parse_transactions_from_excel(df)
    
    # Otherwise, parse from text (PDF)
    lines = text.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        i += 1
        
        if not line:
            continue
        
        # Match transaction pattern: DD/MM/YY at start
        date_match = re.match(r'^(\d{2}/\d{2}/\d{2})', line)
        if not date_match:
            continue
        
        # Extract amounts (numbers with decimals)
        amounts = re.findall(r'[\d,]+\.\d{2}', line)
        if len(amounts) < 2:
            continue
        
        try:
            # Parse amounts
            transaction_amount = float(amounts[-2].replace(',', ''))
            current_balance = float(amounts[-1].replace(',', ''))
            
            # Extract complete description
            description = re.sub(r'^\d{2}/\d{2}/\d{2}\s+', '', line)
            trans_amount_str = amounts[-2]
            amount_index = description.rfind(trans_amount_str)
            if amount_index != -1:
                description = description[:amount_index].strip()
            
            # Now check if next lines are continuation lines (don't start with date, no amounts)
            while i < len(lines):
                next_line = lines[i].strip()
                if not next_line:  # Skip empty lines
                    i += 1
                    continue
                
                # Stop if next line starts with a date (new transaction)
                if re.match(r'^\d{2}/\d{2}/\d{2}', next_line):
                    break
                
                # Stop if next line has significant amounts (likely a new transaction)
                next_amounts = re.findall(r'[\d,]+\.\d{2}', next_line)
                if next_amounts and len(next_amounts) >= 2:
                    break
                
                # This is a continuation line - add it to description
                description = description + " " + next_line
                i += 1
            
            # Clean up spaces
            description = re.sub(r'\s+', ' ', description).strip()
            
            # Determine transaction type by balance change
            transaction_type = "Unknown"
            if previous_balance is not None:
                balance_change = current_balance - previous_balance
                if abs(balance_change - transaction_amount) < 0.01:
                    transaction_type = "Credit"  # Balance increased
                elif abs(balance_change + transaction_amount) < 0.01:
                    transaction_type = "Debit"   # Balance decreased
            else:
                # For first transaction, check narration keywords
                if any(keyword in description.upper() for keyword in ['UPI', 'NEFT', 'RTGS', 'IMPS', 'DEPOSIT', 'SALARY', 'INTEREST', 'CREDIT']):
                    transaction_type = "Credit"
                else:
                    transaction_type = "Debit"
            
            # Extract flat number
            flat_number = extract_flat_number(description)
            
            transaction = {
                'Date': date_match.group(1),
                'Narration': description,
                'Amount': transaction_amount,
                'Balance': current_balance,
                'Type': transaction_type,
                'Flat_Number': flat_number
            }
            
            transactions.append(transaction)
            previous_balance = current_balance
            
        except (ValueError, IndexError) as e:
            continue
    
    print(f"✅ Found {len(transactions)} total transactions")
    
    # Separate credits and debits
    credits = [t for t in transactions if t['Type'] == 'Credit']
    debits = [t for t in transactions if t['Type'] == 'Debit']
    
    print(f"   💚 Credits: {len(credits)}")
    print(f"   💸 Debits: {len(debits)}")
    print(f"   ❓ Unknown: {len(transactions) - len(credits) - len(debits)}")
    
    return transactions, credits, debits

def parse_transactions_from_excel(df):
    """Parse transactions from Excel dataframe"""
    print("📊 Parsing transactions from Excel...")
    
    transactions = []
    previous_balance = None
    
    # Map column names (case-insensitive)
    columns_lower = {col.lower(): col for col in df.columns}
    
    # Find date column
    date_col = None
    for key in columns_lower:
        if 'date' in key:
            date_col = columns_lower[key]
            break
    
    # Find narration column
    narration_col = None
    for key in columns_lower:
        if 'narration' in key or 'description' in key:
            narration_col = columns_lower[key]
            break
    
    # Find deposit/credit column
    deposit_col = None
    for key in columns_lower:
        if 'deposit' in key or 'credit' in key.lower():
            deposit_col = columns_lower[key]
            break
    
    # Find withdrawal/debit column
    withdrawal_col = None
    for key in columns_lower:
        if 'withdrawal' in key or 'debit' in key.lower():
            withdrawal_col = columns_lower[key]
            break
    
    # Find balance column
    balance_col = None
    for key in columns_lower:
        if 'balance' in key:
            balance_col = columns_lower[key]
            break
    
    print(f"Detected columns: Date={date_col}, Narration={narration_col}, Deposit={deposit_col}, Withdrawal={withdrawal_col}, Balance={balance_col}")
    
    # Process each row
    for idx, row in df.iterrows():
        try:
            # Skip empty rows
            if pd.isna(row.get(date_col, "")):
                continue
            
            # Extract date
            date_val = row.get(date_col, "")
            date_str = str(date_val).strip()
            
            # Try to convert datetime to DD/MM/YY format
            if isinstance(date_val, pd.Timestamp):
                date_str = date_val.strftime("%d/%m/%y")
            
            # Extract narration
            narration = str(row.get(narration_col, "")).strip() if narration_col else ""
            
            # Extract deposit/credit amount
            deposit_amount = 0
            if deposit_col and pd.notna(row.get(deposit_col)):
                try:
                    deposit_amount = float(str(row.get(deposit_col, "0")).replace(',', ''))
                except (ValueError, TypeError):
                    deposit_amount = 0
            
            # Extract withdrawal/debit amount
            withdrawal_amount = 0
            if withdrawal_col and pd.notna(row.get(withdrawal_col)):
                try:
                    withdrawal_amount = float(str(row.get(withdrawal_col, "0")).replace(',', ''))
                except (ValueError, TypeError):
                    withdrawal_amount = 0
            
            # Determine amount and type
            if deposit_amount > 0:
                transaction_amount = deposit_amount
                transaction_type = "Credit"
            elif withdrawal_amount > 0:
                transaction_amount = withdrawal_amount
                transaction_type = "Debit"
            else:
                continue
            
            # Extract balance
            balance = 0
            if balance_col and pd.notna(row.get(balance_col)):
                try:
                    balance = float(str(row.get(balance_col, "0")).replace(',', ''))
                except (ValueError, TypeError):
                    balance = 0
            
            # Extract flat number
            flat_number = extract_flat_number(narration)
            
            transaction = {
                'Date': date_str,
                'Narration': narration,
                'Amount': transaction_amount,
                'Balance': balance,
                'Type': transaction_type,
                'Flat_Number': flat_number
            }
            
            transactions.append(transaction)
            previous_balance = balance
            
        except Exception as e:
            continue
    
    print(f"✅ Found {len(transactions)} transactions from Excel")
    
    # Separate credits and debits
    credits = [t for t in transactions if t['Type'] == 'Credit']
    debits = [t for t in transactions if t['Type'] == 'Debit']
    
    print(f"   💚 Credits: {len(credits)}")
    print(f"   💸 Debits: {len(debits)}")
    print(f"   ❓ Unknown: {len(transactions) - len(credits) - len(debits)}")
    
    return transactions, credits, debits

def organize_flat_payments(credits):
    """Organize payments by flat number with each payment in separate columns"""
    print("\n🏠 STEP 2B: Organizing Payments by Flat")
    print("=" * 40)
    
    from collections import defaultdict
    
    # Group payments by flat number (only include actual flats, not "Unknown" or "TOTAL CREDITS")
    flat_data = defaultdict(lambda: {'payments': [], 'dates': [], 'narrations': []})
    
    for credit in credits:
        flat = credit.get('Flat_Number')
        amount = credit.get('Amount')
        
        # Validate flat number and amount
        if not flat or flat in ['Unknown', 'TOTAL CREDITS', 'TOTAL DEBITS', '']:
            continue
        
        # Skip if amount is None, NaN, or not a valid number
        if amount is None or (isinstance(amount, float) and pd.isna(amount)):
            continue
        
        try:
            # Ensure amount is numeric
            float(amount)
        except (ValueError, TypeError):
            continue
        
        flat_data[flat]['payments'].append(amount)
        flat_data[flat]['dates'].append(credit.get('Date', ''))
        flat_data[flat]['narrations'].append(credit.get('Narration', ''))
    
    # Convert to list for processing, sorted by flat number
    flat_details = []
    for flat_num in sorted(flat_data.keys()):
        data = flat_data[flat_num]
        if data['payments']:  # Only include flats with at least one payment
            flat_details.append({
                'Flat_Number': flat_num,
                'Payment_Count': len(data['payments']),
                'Payments': data['payments'],
                'Dates': data['dates'],
                'Narrations': data['narrations']
            })
    
    # Count flats with multiple payments
    multi_payment_flats = [f for f in flat_details if f['Payment_Count'] > 1]
    print(f"✅ Found {len(multi_payment_flats)} flats with multiple payments")
    print(f"   Total flats with payments: {len(flat_details)}")
    
    return flat_details

def create_comprehensive_excel(transactions, credits, debits):
    """Create Excel with Credits and Debits in separate sheets"""
    print("\n📊 STEP 3: Creating Comprehensive Excel File")
    print("=" * 45)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook with multiple sheets
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create credits sheet
        if credits:
            credits_df = pd.DataFrame(credits)
            credits_ws = wb.create_sheet("Credits")
            
            # Add headers
            headers = ['Date', 'Narration', 'Amount (₹)', 'Flat Number']
            credits_ws.append(headers)
            
            # Style headers
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal="center")
            
            for col in range(1, len(headers) + 1):
                cell = credits_ws.cell(row=1, column=col)
                cell.fill = yellow_fill
                cell.font = bold_font
                cell.alignment = center_alignment
            
            # Add data
            for idx, credit in enumerate(credits, start=2):
                credits_ws.cell(row=idx, column=1, value=credit['Date'])
                credits_ws.cell(row=idx, column=2, value=credit['Narration'])
                credits_ws.cell(row=idx, column=3, value=credit['Amount'])
                credits_ws.cell(row=idx, column=4, value=credit['Flat_Number'] or '')
                
                # Highlight rows with flat numbers
                if credit['Flat_Number']:
                    for col in range(1, 5):
                        credits_ws.cell(row=idx, column=col).fill = PatternFill(
                            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
                        )
            
            # Add total credits row
            total_row = len(credits) + 2
            credits_ws.cell(row=total_row, column=1, value="TOTAL CREDITS")
            credits_ws.cell(row=total_row, column=2, value="")
            credits_ws.cell(row=total_row, column=3, value=sum(c['Amount'] for c in credits))
            
            # Style total row
            total_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            for col in range(1, 5):
                cell = credits_ws.cell(row=total_row, column=col)
                cell.fill = total_fill
                cell.font = Font(bold=True)
            
            # Auto-size columns
            for col in credits_ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                credits_ws.column_dimensions[column].width = adjusted_width
        
        # Create debits sheet
        if debits:
            debits_df = pd.DataFrame(debits)
            debits_ws = wb.create_sheet("Debits")
            
            # Add headers
            headers = ['Date', 'Narration', 'Amount (₹)']
            debits_ws.append(headers)
            
            # Style headers
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            for col in range(1, len(headers) + 1):
                cell = debits_ws.cell(row=1, column=col)
                cell.fill = red_fill
                cell.font = bold_font
                cell.alignment = center_alignment
            
            # Add data
            for idx, debit in enumerate(debits, start=2):
                debits_ws.cell(row=idx, column=1, value=debit['Date'])
                debits_ws.cell(row=idx, column=2, value=debit['Narration'])
                debits_ws.cell(row=idx, column=3, value=debit['Amount'])
            
            # Add total debits row
            total_row = len(debits) + 2
            debits_ws.cell(row=total_row, column=1, value="TOTAL DEBITS")
            debits_ws.cell(row=total_row, column=2, value="")
            debits_ws.cell(row=total_row, column=3, value=sum(d['Amount'] for d in debits))
            
            # Style total row
            total_fill = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")
            for col in range(1, 4):
                cell = debits_ws.cell(row=total_row, column=col)
                cell.fill = total_fill
                cell.font = Font(bold=True)
            
            # Auto-size columns
            for col in debits_ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                debits_ws.column_dimensions[column].width = adjusted_width
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary", 0)  # Insert at beginning
        
        # Summary data
        credits_with_flats = [c for c in credits if c['Flat_Number']]
        debits_with_flats = [d for d in debits if d['Flat_Number']]
        
        total_credits_amount = sum(c['Amount'] for c in credits)
        total_debits_amount = sum(d['Amount'] for d in debits)
        
        # Add summary information
        summary_data = [
            ["HDFC Bank Statement Analysis", ""],
            ["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["TRANSACTION SUMMARY", ""],
            ["Total Credits:", f"{len(credits)} transactions"],
            ["Total Credit Amount:", f"₹{total_credits_amount:,.2f}"],
            ["Credits with Flat Numbers:", f"{len(credits_with_flats)} transactions"],
            ["", ""],
            ["Total Debits:", f"{len(debits)} transactions"],
            ["Total Debit Amount:", f"₹{total_debits_amount:,.2f}"],
            ["Debits with Flat Numbers:", f"{len(debits_with_flats)} transactions"],
            ["", ""],
            ["Net Amount:", f"₹{total_credits_amount - total_debits_amount:,.2f}"],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_idx, column=2, value=value)
        
        # Add flat breakdown if any
        if credits_with_flats or debits_with_flats:
            summary_ws.cell(row=len(summary_data) + 2, column=1, value="FLAT NUMBER BREAKDOWN").font = Font(bold=True)
            
            # Credits by flat
            if credits_with_flats:
                credits_by_flat = {}
                for credit in credits_with_flats:
                    flat = credit['Flat_Number']
                    if flat not in credits_by_flat:
                        credits_by_flat[flat] = []
                    credits_by_flat[flat].append(credit['Amount'])
                
                row_idx = len(summary_data) + 4
                summary_ws.cell(row=row_idx, column=1, value="Credits by Flat:").font = Font(bold=True)
                row_idx += 1
                
                for flat, amounts in sorted(credits_by_flat.items()):
                    summary_ws.cell(row=row_idx, column=1, value=f"  {flat}:")
                    summary_ws.cell(row=row_idx, column=2, value=f"₹{sum(amounts):,.2f} ({len(amounts)} transactions)")
                    row_idx += 1
            
            # Debits by flat
            if debits_with_flats:
                debits_by_flat = {}
                for debit in debits_with_flats:
                    flat = debit['Flat_Number']
                    if flat not in debits_by_flat:
                        debits_by_flat[flat] = []
                    debits_by_flat[flat].append(debit['Amount'])
                
                row_idx += 1
                summary_ws.cell(row=row_idx, column=1, value="Debits by Flat:").font = Font(bold=True)
                row_idx += 1
                
                for flat, amounts in sorted(debits_by_flat.items()):
                    summary_ws.cell(row=row_idx, column=1, value=f"  {flat}:")
                    summary_ws.cell(row=row_idx, column=2, value=f"₹{sum(amounts):,.2f} ({len(amounts)} transactions)")
                    row_idx += 1
        
        # Auto-size summary columns
        for col in summary_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            summary_ws.column_dimensions[column].width = adjusted_width
        
        # Add sheet with only entries that have flat numbers (one row per transaction, flats-first, include flats with no transactions)
        # Build flat list in desired order (same as flat details ranges)
        flats = []
        # A blocks
        flats += [f"A{str(i).zfill(3)}" for i in range(1, 21)]
        flats += [f"A{str(i).zfill(3)}" for i in range(101, 121)]
        flats += [f"A{str(i).zfill(3)}" for i in range(201, 221)]
        flats += [f"A{str(i).zfill(3)}" for i in range(301, 321)]
        # B blocks
        flats += [f"B{str(i).zfill(3)}" for i in range(1, 13)]
        flats += [f"B{str(i).zfill(3)}" for i in range(101, 113)]
        flats += [f"B{str(i).zfill(3)}" for i in range(201, 213)]
        flats += [f"B{str(i).zfill(3)}" for i in range(301, 313)]
        # C blocks
        flats += [f"C{str(i).zfill(3)}" for i in range(1, 19)]
        flats += [f"C{str(i).zfill(3)}" for i in range(101, 119)]
        flats += [f"C{str(i).zfill(3)}" for i in range(201, 219)]
        flats += [f"C{str(i).zfill(3)}" for i in range(301, 319)]

        # Group transactions by normalized flat
        flat_entries_map = {f: [] for f in flats}
        for t in transactions:
            flat_raw = str(t.get('Flat_Number') or '').upper().strip()
            flat_norm = re.sub(r"\s+", "", flat_raw)
            if flat_norm in flat_entries_map:
                flat_entries_map[flat_norm].append(t)

        # Create sheet (one row per flat, payments in separate columns)
        entries_ws = wb.create_sheet("Entries with Flats")

        # Build payments per flat (list of numeric amounts)
        flat_payments_map = {f: [] for f in flats}
        for t in transactions:
            flat_raw = str(t.get('Flat_Number') or '').upper().strip()
            flat_norm = re.sub(r"\s+", "", flat_raw)
            if flat_norm in flat_payments_map:
                try:
                    amt = float(t.get('Amount') or 0)
                    flat_payments_map[flat_norm].append(amt)
                except (ValueError, TypeError):
                    # skip non-numeric amounts
                    pass

        # Determine maximum number of payments across all flats to build columns
        max_payments = max((len(v) for v in flat_payments_map.values()), default=0)

        # Prepare headers: Flat Number | Payment 1 | Payment 2 | ... | Payment N | Total | Payment Count
        headers = ['Flat Number'] + [f'Payment {i}' for i in range(1, max_payments + 1)] + ['Total', 'Payment Count']
        entries_ws.append(headers)

        # Style headers
        header_fill2 = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = entries_ws.cell(row=1, column=col)
            cell.fill = header_fill2
            cell.font = bold_font
            cell.alignment = center_alignment

        # Populate one row per flat in the defined order, with placeholders for no payments
        row_idx = 2
        for flat in flats:
            payments = flat_payments_map.get(flat, [])
            entries_ws.cell(row=row_idx, column=1, value=flat)

            # Fill payments into columns
            for p_idx in range(max_payments):
                col_idx = 2 + p_idx
                if p_idx < len(payments):
                    entries_ws.cell(row=row_idx, column=col_idx, value=payments[p_idx])
                else:
                    # leave blank for missing payments
                    entries_ws.cell(row=row_idx, column=col_idx, value=None)

            # Total for the flat (sum of payment columns)
            total_col_idx = 2 + max_payments
            start_col_letter = entries_ws.cell(row=1, column=2).column_letter
            end_col_letter = entries_ws.cell(row=1, column=1 + max_payments).column_letter if max_payments >= 1 else start_col_letter
            if max_payments >= 1:
                total_formula = f"=SUM({start_col_letter}{row_idx}:{end_col_letter}{row_idx})"
                entries_ws.cell(row=row_idx, column=total_col_idx, value=total_formula)
            else:
                entries_ws.cell(row=row_idx, column=total_col_idx, value=0)

            # Payment count
            count_col_idx = total_col_idx + 1
            entries_ws.cell(row=row_idx, column=count_col_idx, value=len(payments))

            row_idx += 1

        # Add total row (sum of each payment column, and grand total)
        total_row_idx = row_idx + 0
        entries_ws.cell(row=total_row_idx, column=1, value="TOTAL (Flats)")
        # Sum each payment column
        for p_idx in range(max_payments):
            col_idx = 2 + p_idx
            col_letter = entries_ws.cell(row=1, column=col_idx).column_letter
            sum_formula = f"=SUM({col_letter}2:{col_letter}{row_idx - 1})"
            cell = entries_ws.cell(row=total_row_idx, column=col_idx, value=sum_formula)
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Grand total column
        total_col_idx = 2 + max_payments
        total_col_letter = entries_ws.cell(row=1, column=total_col_idx).column_letter
        grand_total_formula = f"=SUM({total_col_letter}2:{total_col_letter}{row_idx - 1})"
        cell = entries_ws.cell(row=total_row_idx, column=total_col_idx, value=grand_total_formula)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        # Payment count total (optional) - sum of counts
        count_col_idx = total_col_idx + 1
        count_col_letter = entries_ws.cell(row=1, column=count_col_idx).column_letter
        count_sum_formula = f"=SUM({count_col_letter}2:{count_col_letter}{row_idx - 1})"
        cell = entries_ws.cell(row=total_row_idx, column=count_col_idx, value=count_sum_formula)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Style total row label
        entries_ws.cell(row=total_row_idx, column=1).font = Font(bold=True)
        entries_ws.cell(row=total_row_idx, column=1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Auto-size columns
        for col in entries_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            entries_ws.column_dimensions[column].width = adjusted_width

        # Add Flat Details sheet
        flat_details = organize_flat_payments(credits)
        if flat_details:
            add_flat_details_sheet(wb, flat_details)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = Path("output") / f"HDFC_Complete_Statement_{timestamp}.xlsx"
        output_file.parent.mkdir(exist_ok=True)
        
        wb.save(output_file)

        print(f"✅ Excel file saved: {output_file}")
        print(f"📊 Contains {len(wb.sheetnames)} sheets:")
        for sheet_name in wb.sheetnames:
            print(f"   📋 {sheet_name}")

        # Export each sheet to PDF (optional)
        try:
            save_sheets_as_pdfs(output_file)
        except Exception:
            pass

        return output_file
        
    except Exception as e:
        print(f"❌ Error creating Excel: {e}")
        return None

def add_flat_details_sheet(wb, flat_details):
    """Add a sheet with flat-wise payment breakdown"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    if not flat_details:
        return
    
    # Filter out invalid flat entries and find max number of payments
    valid_flats = [f for f in flat_details if f.get('Flat_Number') and str(f.get('Flat_Number')).strip() and f.get('Payment_Count', 0) > 0]
    
    if not valid_flats:
        return
    
    max_payments = max([f['Payment_Count'] for f in valid_flats])
    
    # Create new sheet
    flat_sheet = wb.create_sheet("Flat Details")
    
    # Create headers: Flat Number | Payment 1 | Payment 2 | ... | Payment N | Total
    headers = ['Flat Number']
    for i in range(1, max_payments + 1):
        headers.append(f'Payment {i}')
    headers.append('Total')
    headers.append('Payment Count')
    
    flat_sheet.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = flat_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Add data rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_idx, flat in enumerate(valid_flats, start=2):
        col = 1
        
        # Flat number
        cell = flat_sheet.cell(row=row_idx, column=col, value=flat['Flat_Number'])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        col += 1
        
        # Payments - only include numeric values
        for payment in flat['Payments']:
            try:
                payment_val = float(payment) if payment else 0
                cell = flat_sheet.cell(row=row_idx, column=col, value=payment_val)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            except (ValueError, TypeError):
                # Skip non-numeric payments
                pass
            col += 1
        
        # Empty cells for missing payments
        for _ in range(max_payments - len(flat['Payments'])):
            col += 1
        
        # Total - calculate only from numeric values
        numeric_payments = []
        for p in flat['Payments']:
            try:
                numeric_payments.append(float(p) if p else 0)
            except (ValueError, TypeError):
                pass
        
        total = sum(numeric_payments)
        cell = flat_sheet.cell(row=row_idx, column=col, value=total)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal="right")
        col += 1
        
        # Payment count
        cell = flat_sheet.cell(row=row_idx, column=col, value=flat['Payment_Count'])
        cell.alignment = Alignment(horizontal="center")
    
    # Add total row
    total_row = len(valid_flats) + 2
    flat_sheet.cell(row=total_row, column=1, value="TOTAL ALL FLATS")
    flat_sheet.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Calculate column totals only for payment columns (not total or count)
    for col_idx in range(2, max_payments + 2):
        col_letter = flat_sheet.cell(row=1, column=col_idx).column_letter
        total_formula = f'=SUM({col_letter}2:{col_letter}{len(valid_flats) + 1})'
        cell = flat_sheet.cell(row=total_row, column=col_idx, value=total_formula)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Total amount column - sum of all totals
    total_col = max_payments + 2
    total_col_letter = flat_sheet.cell(row=1, column=total_col).column_letter
    grand_total_formula = f'=SUM({total_col_letter}2:{total_col_letter}{len(valid_flats) + 1})'
    cell = flat_sheet.cell(row=total_row, column=total_col, value=grand_total_formula)
    cell.number_format = '#,##0.00'
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Auto-size columns
    for col_idx in range(1, max_payments + 4):
        max_length = 12
        col_letter = flat_sheet.cell(row=1, column=col_idx).column_letter
        for cell in flat_sheet[f'{col_letter}1:{col_letter}{len(valid_flats) + 2}']:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        flat_sheet.column_dimensions[col_letter].width = adjusted_width

def save_sheets_as_pdfs(excel_path):
    """Export each sheet in the given Excel workbook to a separate PDF file.
    Tries to use reportlab for high-quality, colored PDFs (landscape, zebra stripes, header styling).
    If reportlab cannot be imported or installed, falls back to generating a colorized HTML file per sheet for visual inspection.
    """
    # Local import helpers
    try:
        from openpyxl import load_workbook
    except Exception:
        print("❌ openpyxl is required for PDF/HTML export. Please install it: pip install openpyxl")
        return

    # Try to import reportlab; set flag whether we can produce PDF directly
    can_pdf = True
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
    except Exception:
        can_pdf = False

    if not can_pdf:
        # Try auto-install reportlab once
        try:
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'reportlab'])
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            can_pdf = True
        except Exception:
            can_pdf = False

    wb = load_workbook(excel_path, data_only=True)
    out_dir = Path(excel_path).parent

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([('' if c is None else c) for c in row])

        if not rows:
            continue

        safe_sheet = re.sub(r"[^A-Za-z0-9]+", "_", sheet_name).strip('_')
        base_name = Path(excel_path).stem
        pdf_path = out_dir / f"{base_name}_{safe_sheet}.pdf"
        html_path = out_dir / f"{base_name}_{safe_sheet}.html"

        # Determine numeric/date/narration columns from header
        header_texts = [str(c).lower() if c is not None else '' for c in rows[0]]
        amount_col_idx = None
        date_col_idx = None
        narration_col_idx = None
        for idx, ht in enumerate(header_texts):
            if 'amount' in ht or 'amt' in ht:
                amount_col_idx = idx
            if 'date' in ht:
                date_col_idx = idx
            if 'narr' in ht or 'descrip' in ht or 'narration' in ht or 'description' in ht:
                narration_col_idx = idx

        if can_pdf:
            # Build a nicely styled PDF using reportlab in landscape for readability
            try:
                # Styling
                page_size = landscape(A4)
                left_margin = right_margin = 12 * mm
                top_margin = bottom_margin = 12 * mm

                styles = getSampleStyleSheet()
                header_style = ParagraphStyle('header', parent=styles['Heading2'], alignment=1, fontSize=12, leading=14)
                small_bold = ParagraphStyle('small_bold', parent=styles['Normal'], alignment=1, fontSize=8, leading=10)
                body_style = ParagraphStyle('body', parent=styles['BodyText'], fontSize=8, leading=10)
                amt_style = ParagraphStyle('amt', parent=styles['BodyText'], fontSize=8, leading=10, alignment=2)
                narr_style = ParagraphStyle('narr', parent=styles['BodyText'], fontSize=8, leading=10)

                # Prepare table data with Paragraphs to allow wrapping
                num_cols = max(len(r) for r in rows)
                formatted_rows = []
                # Header row as bold
                hdr = []
                for c_idx in range(num_cols):
                    txt = rows[0][c_idx] if c_idx < len(rows[0]) else ''
                    hdr.append(Paragraph(f"<b>{str(txt)}</b>", small_bold))
                formatted_rows.append(hdr)

                for r_idx, r in enumerate(rows[1:], start=1):
                    new_row = []
                    for c_idx in range(num_cols):
                        cell = ''
                        if c_idx < len(r) and r[c_idx] is not None:
                            cell = r[c_idx]
                        # Format amount
                        if amount_col_idx is not None and c_idx == amount_col_idx and cell != '':
                            try:
                                num = float(cell)
                                txt = f"{num:,.2f}"
                                new_row.append(Paragraph(txt, amt_style))
                            except Exception:
                                new_row.append(Paragraph(str(cell), body_style))
                        elif date_col_idx is not None and c_idx == date_col_idx and cell != '':
                            new_row.append(Paragraph(str(cell), ParagraphStyle('date', parent=body_style, alignment=1)))
                        elif c_idx == narration_col_idx:
                            new_row.append(Paragraph(str(cell).replace('\n', '<br/>'), narr_style))
                        else:
                            new_row.append(Paragraph(str(cell), body_style))
                    formatted_rows.append(new_row)

                # Compute column widths proportionally, boost narration
                # Measure max characters per column
                col_max_chars = [0] * num_cols
                for r in rows:
                    for i, cell in enumerate(r):
                        try:
                            text = str(cell)
                        except Exception:
                            text = ''
                        if len(text) > col_max_chars[i]:
                            col_max_chars[i] = len(text)

                # Build weights: give narration extra room
                weights = []
                for i, chars in enumerate(col_max_chars):
                    base = max(chars, 1)
                    if narration_col_idx is not None and i == narration_col_idx:
                        weight = base * 4.0
                    elif amount_col_idx is not None and i == amount_col_idx:
                        weight = base * 0.9
                    elif date_col_idx is not None and i == date_col_idx:
                        weight = base * 0.9
                    else:
                        weight = base * 1.0
                    weights.append(weight)

                total_w = sum(weights) if sum(weights) > 0 else 1
                usable_width = page_size[0] - left_margin - right_margin

                # Compute proportional widths then apply min/max caps and normalize to usable_width
                provisional = [usable_width * (w / total_w) for w in weights]
                col_widths = []
                for i, p in enumerate(provisional):
                    # sensible minimums: narration larger
                    if narration_col_idx is not None and i == narration_col_idx:
                        min_w = 45 * mm
                        max_w = usable_width * 0.8
                    elif amount_col_idx is not None and i == amount_col_idx:
                        min_w = 20 * mm
                        max_w = usable_width * 0.3
                    elif date_col_idx is not None and i == date_col_idx:
                        min_w = 22 * mm
                        max_w = usable_width * 0.25
                    else:
                        min_w = 25 * mm
                        max_w = usable_width * 0.5

                    w = max(min_w, min(p, max_w))
                    col_widths.append(w)

                total_width = sum(col_widths)
                if total_width == 0:
                    # fallback to equal columns
                    col_widths = [usable_width / num_cols] * num_cols
                elif total_width > usable_width:
                    # normalize down to usable_width
                    ratio = usable_width / total_width
                    col_widths = [w * ratio for w in col_widths]
                elif total_width < usable_width and narration_col_idx is not None:
                    # if not full, give remaining space to narration column to improve readability
                    extra = usable_width - total_width
                    col_widths[narration_col_idx] += extra

                 # Build table and apply zebra styling
                table = Table(formatted_rows, colWidths=col_widths, repeatRows=1)
                tbl_style = TableStyle([
                    ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#BBBBBB')),
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#FFD966')),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ])

                # Add zebra stripes
                for zi in range(1, len(formatted_rows)):
                    bg = colors.whitesmoke if zi % 2 == 0 else colors.white
                    tbl_style.add('BACKGROUND', (0, zi), (-1, zi), bg)

                # Right align amount and center date
                if amount_col_idx is not None:
                    tbl_style.add('ALIGN', (amount_col_idx, 1), (amount_col_idx, -1), 'RIGHT')
                if date_col_idx is not None:
                    tbl_style.add('ALIGN', (date_col_idx, 1), (date_col_idx, -1), 'CENTER')

                # Emphasize totals row if last row header contains TOTAL
                last_row_idx = len(formatted_rows) - 1
                try:
                    last_vals = [r.text for r in formatted_rows[-1]]
                except Exception:
                    last_vals = []
                # If the sheet's last row label contains TOTAL or TOTAL (Flats)
                if any('TOTAL' in str(v).upper() for v in rows[-1]):
                    tbl_style.add('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor('#C6EFCE'))
                    tbl_style.add('FONTNAME', (0, last_row_idx), (-1, last_row_idx), 'Helvetica-Bold')

                table.setStyle(tbl_style)

                # Build document
                doc = SimpleDocTemplate(str(pdf_path), pagesize=page_size, leftMargin=left_margin, rightMargin=right_margin, topMargin=top_margin, bottomMargin=bottom_margin)
                elements = []
                # Header
                title = f"{base_name} - {sheet_name}"
                elements.append(Paragraph(title, header_style))
                elements.append(Spacer(1, 4 * mm))
                elements.append(table)

                doc.build(elements)
                print(f"✅ Saved PDF for sheet '{sheet_name}': {pdf_path}")
                continue
            except Exception as e:
                print(f"❌ Failed to create PDF for sheet '{sheet_name}' using reportlab: {e}")
                # fall through to HTML fallback

        # Fallback: create colorized HTML representation
        try:
            css = '''
                @page { size: A4 landscape; margin: 10mm; }
                body { font-family: Arial, Helvetica, sans-serif; font-size:12px; margin:0; padding:0; }
                .header { background:#FFD966; padding:8px; text-align:center; font-weight:bold; }
                table { border-collapse: collapse; width:100%; table-layout: fixed; }
                th, td { border: 1px solid #ccc; padding:6px; vertical-align:top; word-wrap: break-word; }
                th { background: #FFD966; text-align:center; }
                tr:nth-child(even) td { background: #f9f9f9; }
                .amt { text-align:right; font-variant-numeric: tabular-nums; }
                .center { text-align:center; }
                /* make narration column wider for readability */
                .narr { width: 45%; }
            '''
            html_parts = []
            html_parts.append(f"<html><head><meta charset='utf-8'><title>{base_name} - {sheet_name}</title><style>{css}</style></head><body>")
            html_parts.append(f"<div class='header'>{base_name} &nbsp; - &nbsp; {sheet_name}</div>")
            html_parts.append('<table>')

            # Header row
            html_parts.append('<tr>')
            for c in rows[0]:
                html_parts.append(f"<th>{str(c)}</th>")
            html_parts.append('</tr>')

            # Data rows
            for r in rows[1:]:
                html_parts.append('<tr>')
                for i, c in enumerate(r):
                    cell_html = '' if c is None else str(c)
                    # format amount
                    if amount_col_idx is not None and i == amount_col_idx and cell_html != '':
                        try:
                            num = float(c)
                            cell_html = f"{num:,.2f}"
                            html_parts.append(f"<td class='amt'>{cell_html}</td>")
                            continue
                        except Exception:
                            pass
                    # prepare escaped HTML for cell (replace newlines)
                    cell_html_escaped = cell_html.replace('\n', '<br/>')
                    if date_col_idx is not None and i == date_col_idx:
                        html_parts.append(f"<td class='center'>{cell_html_escaped}</td>")
                    else:
                        html_parts.append(f"<td>{cell_html_escaped}</td>")
                html_parts.append('</tr>')

            html_parts.append('</table>')
            html_parts.append(f"<p style='font-size:11px;color:#555;margin-top:8px;'>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>")
            html_parts.append('</body></html>')

            with open(html_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(html_parts))

            print(f"ℹ️ ReportLab not available — created HTML fallback for sheet '{sheet_name}': {html_path}")
            print("👉 To generate PDFs, install reportlab: python3 -m pip install reportlab")
        except Exception as e:
            print(f"❌ Failed to create HTML fallback for sheet '{sheet_name}': {e}")

def generate_flat_config(transactions, out_path=None):
    """Generate flat configuration JSON from transaction data.
    Heuristic: if narration contains a flat number (detected earlier), use that; else use common name tokens.
    
    Args:
      transactions: list of transaction dicts (from parse_transactions)
      out_path: optional output path for flatConfig.json
    
    Returns:
      Path to written JSON file, or None if failed.
    """
    from collections import defaultdict
    
    # Heuristic: build mapping from flat number <-> common name tokens
    mapping = []
    used_flats = set()  # track used flat numbers to avoid duplicates
    for t in transactions:
        flat = t.get('Flat_Number')
        if not flat or flat in ['Unknown', 'TOTAL CREDITS', 'TOTAL DEBITS']:
            continue
        if flat in used_flats:
            continue
        used_flats.add(flat)
        
        # Extract common name tokens from narration
        narr = (t.get('Narration') or '').upper()
        narr_tokens = re.findall(r'\b[A-Z]{1,3}\d{1,3}\b', narr)
        narr_tokens = list(sorted(set(narr_tokens)))  # unique, sorted
        if narr_tokens:
            primary_name = narr_tokens[0]  # use first token as primary
        else:
            primary_name = narr  # fallback to full narration
        
        mapping.append({
            'Flat_Number': flat,
            'Name': primary_name,
            'Other_Names': [n for n in narr_tokens[1:]],  # rest as other possible names
        })
    
    # Write to JSON file
    if out_path is None:
        out_path = Path('config') / 'flat_config.json'
    try:
        with open(out_path, 'w', encoding='utf-8') as fh:
            json.dump(mapping, fh, indent=2, ensure_ascii=False)
        print(f"✅ flatConfig.json written: {out_path} ({len(mapping)} entries)")
        return out_path
    except Exception as e:
        print(f"❌ Failed to write flatConfig.json: {e}")
        return None

def load_flat_details_config(path=None):
    """Load config/flat_details.json and return list of entries.

    Returns list of dicts with keys 'Flat_Number' and 'Name'.
    """
    if path is None:
        path = Path('config') / 'flat_details.json'
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            data = json.load(fh)
            if isinstance(data, list):
                return data
    except FileNotFoundError:
        print(f"ℹ️ flat details config not found at {path}")
    except Exception as e:
        print(f"❌ Failed reading flat details config: {e}")
    return []

def assign_flats_from_flat_details(transactions, config_path=None):
    """For transactions missing Flat_Number, try to assign from flat_details.json using name matching.

    Heuristic: if the configured Name (full or tokens) appears in the narration, assign that Flat_Number.
    Returns number of assignments made.
    """
    config = load_flat_details_config(config_path)
    if not config:
        return 0

    # Build normalized name -> flat mapping (allow multiple names mapping to same flat)
    candidates = []
    for entry in config:
        name = (entry.get('Name') or '').strip()
        flat = entry.get('Flat_Number')
        if name and flat:
            # normalize name: uppercase, remove punctuation, collapse spaces
            nm = re.sub(r'[^A-Z0-9\s]', ' ', name.upper())
            nm = re.sub(r'\s+', ' ', nm).strip()
            tokens = [t for t in nm.split() if len(t) >= 2]
            if tokens:
                candidates.append((flat, nm, tokens))

    if not candidates:
        return 0

    assigned = 0
    for t in transactions:
        flat_raw = t.get('Flat_Number')
        if flat_raw and str(flat_raw).strip():
            continue
        narr = (t.get('Narration') or '').upper()
        if not narr:
            continue
        narr_norm = re.sub(r'[^A-Z0-9\s]', ' ', narr)
        narr_norm = re.sub(r'\s+', ' ', narr_norm).strip()

        matched = False
        # Prefer longer/full-name matches first
        candidates_sorted = sorted(candidates, key=lambda x: -len(x[1]))
        for flat, nm, tokens in candidates_sorted:
            # direct substring match
            if nm and nm in narr_norm:
                t['Flat_Number'] = flat
                assigned += 1
                matched = True
                break
            # token subset match
            if tokens and all(tok in narr_norm for tok in tokens):
                t['Flat_Number'] = flat
                assigned += 1
                matched = True
                break
        # no match: continue

    if assigned:
        print(f"✅ Assigned {assigned} missing flat numbers using config/flat_details.json")
    else:
        print("ℹ️ No flat numbers assigned from config")
    return assigned

def main():
    """Main function to process HDFC bank statement"""
    print("🏦 HDFC BANK STATEMENT COMPLETE PROCESSOR")
    print("=" * 50)
    print("🎯 Goal: Extract Credits AND Debits with flat numbers")
    print("📁 Supports: PDF or Excel input")
    print()
    
    # Step 0: Select input source
    input_type, input_file = select_input_source()
    if not input_file:
        print("❌ Failed to select input file")
        return
    
    print("\n📄 STEP 1: Extracting Data")
    print("=" * 35)
    
    # Step 1: Extract data based on source type
    text = None
    df = None
    
    if input_type == "pdf":
        text = extract_from_pdf(input_file)
        if not text:
            print("❌ Failed to extract PDF text")
            return
    
    elif input_type == "excel":
        text, df = extract_from_excel(input_file)
        if text is None or (isinstance(text, pd.DataFrame) and text.empty):
            print("❌ Failed to extract Excel data")
            return
    
    # Step 2: Parse all transactions
    if input_type == "excel":
        all_transactions, credits, debits = parse_transactions_from_excel(df)
    else:
        all_transactions, credits, debits = parse_transactions(text, input_type, df)
    if not all_transactions:
        print("❌ No transactions found")
        return
    
    # Assign flat numbers from config before creating outputs
    assign_flats_from_flat_details(all_transactions)
    
    # Step 3: Create comprehensive Excel
    output_file = create_comprehensive_excel(all_transactions, credits, debits)
    if not output_file:
        print("❌ Failed to create Excel file")
        return

    # Step 4: Save each sheet as PDF
    save_sheets_as_pdfs(output_file)
    
    print("\n" + "="*60)
    print("📋 PROCESSING COMPLETE!")
    print("="*60)
    print(f"✅ Total transactions processed: {len(all_transactions)}")
    print(f"💚 Credits: {len(credits)} transactions")
    print(f"💸 Debits: {len(debits)} transactions")
    
    # Count flat numbers
    credits_with_flats = sum(1 for c in credits if c['Flat_Number'])
    debits_with_flats = sum(1 for d in debits if d['Flat_Number'])
    print(f"🏠 Credits with flat numbers: {credits_with_flats}")
    print(f"🏠 Debits with flat numbers: {debits_with_flats}")
    
    print(f"\n📁 Output file: {output_file}")
    print("🎉 Open the Excel file to view your complete bank statement analysis!")

if __name__ == "__main__":
    main()
